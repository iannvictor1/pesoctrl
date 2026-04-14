from pathlib import Path
from datetime import datetime
import time
import re
import threading
import json
import shutil
import uuid
import sys
import os
import subprocess

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

from config import montar_config, CONFIG_IMPRESSORAS, ARQUIVO_RECEBIMENTO_GLOBAL


EXTENSOES_VALIDAS = {".prn", ".txt", ".zpl", ".epl"}

JANELA_EVENTO_SEGUNDOS = 1.0
MAX_TENTATIVAS = 15
ESPERA_ENTRE_TENTATIVAS = 1.0
LIMPEZA_CACHE_EVENTOS_SEGUNDOS = 30.0

COLUNAS = ["DataHora", "Descricao", "PesoKG", "QtdEtiq", "PesoTotalKG", "Validade"]

CONFIG_LIMPEZA_PADRAO = {
    "dias_reter_etiquetas": 7,
    "dias_reter_pesagens": 30,
    "intervalo_limpeza_automatica_segundos": 600,
}


fila_lock = threading.Lock()
excel_lock = threading.Lock()
log_lock = threading.Lock()
recebimento_lock = threading.Lock()
arquivos_capturados_recentemente = {}


def garantir_pastas(cfg: dict) -> None:
    cfg["base"].mkdir(parents=True, exist_ok=True)
    cfg["pasta_fila"].mkdir(parents=True, exist_ok=True)
    cfg["pasta_historico_etiquetas"].mkdir(parents=True, exist_ok=True)
    cfg["pasta_historico_pesagens"].mkdir(parents=True, exist_ok=True)
    cfg["pasta_descarte_fila"].mkdir(parents=True, exist_ok=True)
    ARQUIVO_RECEBIMENTO_GLOBAL.parent.mkdir(parents=True, exist_ok=True)


def log(cfg: dict, msg: str) -> None:
    linha = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [{cfg['id']}] {msg}"
    with log_lock:
        print(linha)
        with open(cfg["arquivo_log"], "a", encoding="utf-8") as f:
            f.write(linha + "\n")


def caminho_config_limpeza(cfg: dict) -> Path:
    return cfg["base"] / "config_limpeza.json"


def config_limpeza_padrao() -> dict:
    return dict(CONFIG_LIMPEZA_PADRAO)


def salvar_config_limpeza(cfg: dict, dados: dict) -> None:
    caminho = caminho_config_limpeza(cfg)
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def garantir_config_limpeza(cfg: dict) -> None:
    caminho = caminho_config_limpeza(cfg)
    if not caminho.exists():
        salvar_config_limpeza(cfg, config_limpeza_padrao())


def carregar_config_limpeza(cfg: dict) -> dict:
    garantir_config_limpeza(cfg)
    caminho = caminho_config_limpeza(cfg)

    try:
        with open(caminho, "r", encoding="utf-8") as f:
            dados = json.load(f)

        resultado = config_limpeza_padrao()
        resultado.update(dados if isinstance(dados, dict) else {})

        resultado["dias_reter_etiquetas"] = int(resultado.get("dias_reter_etiquetas", 7))
        resultado["dias_reter_pesagens"] = int(resultado.get("dias_reter_pesagens", 30))
        resultado["intervalo_limpeza_automatica_segundos"] = int(
            resultado.get("intervalo_limpeza_automatica_segundos", 600)
        )

        if resultado["dias_reter_etiquetas"] < 0:
            resultado["dias_reter_etiquetas"] = 0
        if resultado["dias_reter_pesagens"] < 0:
            resultado["dias_reter_pesagens"] = 0
        if resultado["intervalo_limpeza_automatica_segundos"] < 10:
            resultado["intervalo_limpeza_automatica_segundos"] = 10

        return resultado

    except Exception:
        dados = config_limpeza_padrao()
        salvar_config_limpeza(cfg, dados)
        return dados


def arquivo_mais_antigo_que(arquivo: Path, dias: int) -> bool:
    try:
        limite = time.time() - (dias * 24 * 60 * 60)
        return arquivo.stat().st_mtime < limite
    except Exception:
        return False


def limpar_historico_pasta(cfg: dict, pasta: Path, dias: int, descricao: str) -> None:
    if dias is None or dias < 0:
        return

    removidos = 0
    erros = 0

    try:
        if not pasta.exists():
            return

        for arquivo in pasta.iterdir():
            if not arquivo.is_file():
                continue

            if arquivo_mais_antigo_que(arquivo, dias):
                try:
                    arquivo.unlink()
                    removidos += 1
                except Exception as e:
                    erros += 1
                    log(cfg, f"Erro ao remover {descricao} antigo {arquivo.name}: {e}")

        if removidos > 0 or erros > 0:
            log(
                cfg,
                f"Limpeza automática em {descricao}: removidos={removidos}, erros={erros}, retenção={dias} dias"
            )
    except Exception as e:
        log(cfg, f"Erro na limpeza automática da pasta {descricao}: {e}")


def executar_limpeza_automatica(cfg: dict) -> None:
    conf = carregar_config_limpeza(cfg)

    limpar_historico_pasta(
        cfg,
        cfg["pasta_historico_etiquetas"],
        conf["dias_reter_etiquetas"],
        "histórico de etiquetas",
    )

    limpar_historico_pasta(
        cfg,
        cfg["pasta_historico_pesagens"],
        conf["dias_reter_pesagens"],
        "histórico de pesagens",
    )


def recebimento_padrao() -> dict:
    return {
        "recebimento_ativo": False,
        "recebimento_id": None,
        "inicio_recebimento": None,
        "contador_pallets": 0,
    }


def salvar_recebimento_global(dados: dict) -> None:
    ARQUIVO_RECEBIMENTO_GLOBAL.parent.mkdir(parents=True, exist_ok=True)
    with open(ARQUIVO_RECEBIMENTO_GLOBAL, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def carregar_recebimento_global() -> dict:
    if not ARQUIVO_RECEBIMENTO_GLOBAL.exists():
        dados = recebimento_padrao()
        salvar_recebimento_global(dados)
        return dados

    try:
        with open(ARQUIVO_RECEBIMENTO_GLOBAL, "r", encoding="utf-8") as f:
            dados = json.load(f)

        base = recebimento_padrao()
        base.update(dados)
        return base
    except Exception:
        dados = recebimento_padrao()
        salvar_recebimento_global(dados)
        return dados


def carregar_recebimento(cfg: dict | None = None) -> dict:
    return carregar_recebimento_global()


def existe_pallet_ativo_em_qualquer_impressora() -> bool:
    for nome_impressora in CONFIG_IMPRESSORAS:
        cfg_tmp = montar_config(nome_impressora)
        garantir_pastas(cfg_tmp)
        garantir_arquivo_status(cfg_tmp)
        status_tmp = carregar_status(cfg_tmp)
        if status_tmp.get("sessao_ativa", False):
            return True
    return False


def iniciar_recebimento_global() -> tuple[bool, str]:
    with recebimento_lock:
        rec = carregar_recebimento_global()

        if rec.get("recebimento_ativo"):
            return False, f"Já existe um recebimento em andamento. ID: {rec.get('recebimento_id')}"

        agora = datetime.now()
        recebimento_id = agora.strftime("%Y%m%d_%H%M%S")

        novo = {
            "recebimento_ativo": True,
            "recebimento_id": recebimento_id,
            "inicio_recebimento": agora.strftime("%Y-%m-%d %H:%M:%S"),
            "contador_pallets": 0,
        }
        salvar_recebimento_global(novo)

    for nome_impressora in CONFIG_IMPRESSORAS:
        cfg_tmp = montar_config(nome_impressora)
        garantir_pastas(cfg_tmp)
        log(cfg_tmp, f"Recebimento global iniciado | ID: {recebimento_id}")

    return True, f"Recebimento iniciado. ID: {recebimento_id}"


def iniciar_recebimento(cfg: dict | None = None) -> tuple[bool, str]:
    return iniciar_recebimento_global()


def encerrar_recebimento_global() -> tuple[bool, str]:
    with recebimento_lock:
        rec = carregar_recebimento_global()

        if not rec.get("recebimento_ativo"):
            return False, "Não há recebimento global ativo para encerrar."

        if existe_pallet_ativo_em_qualquer_impressora():
            return False, "Encerre todos os pallets ativos antes de encerrar o recebimento."

        recebimento_id = rec.get("recebimento_id")
        total_pallets = rec.get("contador_pallets", 0)

        salvar_recebimento_global(recebimento_padrao())

    for nome_impressora in CONFIG_IMPRESSORAS:
        cfg_tmp = montar_config(nome_impressora)
        garantir_pastas(cfg_tmp)
        log(cfg_tmp, f"Recebimento global encerrado | ID: {recebimento_id} | Pallets: {total_pallets}")

    return True, f"Recebimento {recebimento_id} encerrado. Total de pallets: {total_pallets}."


def encerrar_recebimento(cfg: dict | None = None) -> tuple[bool, str]:
    return encerrar_recebimento_global()


def status_padrao() -> dict:
    return {
        "sessao_ativa": False,
        "sessao_id": None,
        "inicio_sessao": None,
        "numero_pallet": None,
        "recebimento_id": None,
        "descricao_produto": None,
    }


def salvar_status(cfg: dict, dados: dict) -> None:
    with open(cfg["arquivo_status"], "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def garantir_arquivo_status(cfg: dict) -> None:
    if not cfg["arquivo_status"].exists():
        salvar_status(cfg, status_padrao())


def carregar_status(cfg: dict) -> dict:
    garantir_arquivo_status(cfg)
    try:
        with open(cfg["arquivo_status"], "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        dados = status_padrao()
        salvar_status(cfg, dados)
        return dados


def sanitizar_para_nome_arquivo(texto: str) -> str:
    texto = re.sub(r"[^\w\s\-]", "", texto, flags=re.UNICODE)
    texto = re.sub(r"\s+", "_", texto.strip())
    return texto[:40] if texto else "sem_descricao"


def montar_nome_excel_pallet(numero_pallet: int, descricao: str, data: str) -> str:
    desc_limpa = sanitizar_para_nome_arquivo(descricao) if descricao else "sem_descricao"
    return f"pallet_{numero_pallet}_{desc_limpa}_{data}.xlsx"


def iniciar_sessao(cfg: dict) -> tuple[bool, str]:
    with recebimento_lock:
        rec = carregar_recebimento_global()

        if not rec.get("recebimento_ativo"):
            return False, "Inicie um recebimento antes de iniciar um pallet."

        status = carregar_status(cfg)
        if status.get("sessao_ativa"):
            return False, f"Já existe um pallet em andamento em {cfg['nome_amigavel']}."

        rec["contador_pallets"] = rec.get("contador_pallets", 0) + 1
        numero_pallet = rec["contador_pallets"]
        salvar_recebimento_global(rec)

    agora = datetime.now()
    sessao_id = agora.strftime("%Y%m%d_%H%M%S")

    novo_status = {
        "sessao_ativa": True,
        "sessao_id": sessao_id,
        "inicio_sessao": agora.strftime("%Y-%m-%d %H:%M:%S"),
        "numero_pallet": numero_pallet,
        "recebimento_id": rec["recebimento_id"],
        "descricao_produto": None,
    }
    salvar_status(cfg, novo_status)

    with excel_lock:
        if not cfg["arquivo_excel_atual"].exists():
            criar_excel_novo(cfg)

    limpar_fila_invalida(cfg)
    executar_limpeza_automatica(cfg)

    log(cfg, f"Pallet iniciado | Recebimento: {rec['recebimento_id']} | Pallet nº {numero_pallet} | Sessão: {sessao_id}")
    return True, f"Pallet {numero_pallet} iniciado em {cfg['nome_amigavel']}."


def encerrar_sessao(cfg: dict) -> tuple[bool, str]:
    status = carregar_status(cfg)

    if not status.get("sessao_ativa"):
        return False, f"Não existe pallet ativo para encerrar em {cfg['nome_amigavel']}."

    sessao_id = status.get("sessao_id") or datetime.now().strftime("%Y%m%d_%H%M%S")
    numero_pallet = status.get("numero_pallet", 1)
    descricao = status.get("descricao_produto") or ""
    data_hoje = datetime.now().strftime("%Y%m%d")

    nome_arquivo = montar_nome_excel_pallet(numero_pallet, descricao, data_hoje)
    destino = cfg["pasta_historico_pesagens"] / nome_arquivo

    contador = 1
    while destino.exists():
        nome_arquivo = montar_nome_excel_pallet(numero_pallet, descricao, data_hoje).replace(".xlsx", f"_{contador}.xlsx")
        destino = cfg["pasta_historico_pesagens"] / nome_arquivo
        contador += 1

    with excel_lock:
        if cfg["arquivo_excel_atual"].exists():
            shutil.move(str(cfg["arquivo_excel_atual"]), str(destino))
            log(cfg, f"Excel do pallet movido para histórico: {destino.name}")

        salvar_status(cfg, status_padrao())
        criar_excel_novo(cfg)

    for caminho in cfg["pasta_fila"].iterdir():
        if not caminho.is_file():
            continue
        if caminho.suffix.lower() not in EXTENSOES_VALIDAS:
            continue

        sessao_arquivo = extrair_sessao_id_do_nome(caminho.name)
        if sessao_arquivo == sessao_id or sessao_arquivo is None:
            mover_fila_para_descarte(cfg, caminho, "encerramento do pallet")

    log(cfg, f"Pallet encerrado | Nº {numero_pallet} | Arquivo: {destino.name}")
    executar_limpeza_automatica(cfg)
    return True, f"Pallet {numero_pallet} encerrado. Arquivo salvo como: {destino.name}"

def novo_pallet(cfg: dict) -> tuple[bool, str]:
    status = carregar_status(cfg)
    
    if not status.get("sessao_ativa"):
        return False, f"Não existe pallet ativo em {cfg['nome_amigavel']} para virar um novo pallet."
    
    ok, msg = encerrar_sessao(cfg)
    if not ok:
        return False, msg
    
    ok, msg2 = iniciar_sessao(cfg)
    if not ok:
        return False, f"Pallet anterior encerrado, mas houve erro ao iniciar o novo: {msg}"
    
    return True, f"Novo pallet iniciado com sucesso em {cfg['nome_amigavel']}."

def criar_excel_novo(cfg: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pesagens"
    ws.append(COLUNAS)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo["A1"] = "Indicador"
    ws_resumo["B1"] = "Valor"
    ws_resumo["A2"] = "Total de etiquetas"
    ws_resumo["A3"] = "Soma total dos pesos"

    for cell in ws_resumo[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_resumo.column_dimensions["A"].width = 25
    ws_resumo.column_dimensions["B"].width = 20

    wb.save(cfg["arquivo_excel_atual"])
    log(cfg, f"Novo Excel criado: {cfg['arquivo_excel_atual'].name}")


def garantir_excel_atual(cfg: dict) -> None:
    if not cfg["arquivo_excel_atual"].exists():
        with excel_lock:
            if not cfg["arquivo_excel_atual"].exists():
                criar_excel_novo(cfg)


def atualizar_resumo_no_workbook(ws_pesagens, ws_resumo) -> None:
    total_etiquetas = 0
    soma_pesos = 0.0

    for row in ws_pesagens.iter_rows(min_row=2, values_only=True):
        _, descricao, peso, qtd_etiquetas, peso_total, validade = row

        if (
            descricao is None
            and peso is None
            and qtd_etiquetas is None
            and peso_total is None
            and validade is None
        ):
            continue

        try:
            if qtd_etiquetas is not None and str(qtd_etiquetas).strip() != "":
                total_etiquetas += int(qtd_etiquetas)
        except Exception:
            pass

        try:
            if peso_total is not None and str(peso_total).strip() != "":
                soma_pesos += float(peso_total)
        except Exception:
            pass

    ws_resumo["B2"] = total_etiquetas
    ws_resumo["B3"] = round(soma_pesos, 3)


def append_registro_excel(cfg: dict, registro: dict) -> None:
    garantir_excel_atual(cfg)

    with excel_lock:
        wb = load_workbook(cfg["arquivo_excel_atual"])
        ws = wb["Pesagens"]
        ws_resumo = wb["Resumo"]

        linha = [registro.get(col, "") for col in COLUNAS]
        ws.append(linha)

        ultima_linha = ws.max_row
        for cell in ws[ultima_linha]:
            cell.alignment = Alignment(vertical="top", wrap_text=False)

        atualizar_resumo_no_workbook(ws, ws_resumo)
        wb.save(cfg["arquivo_excel_atual"])


def limpar_texto_excel(valor):
    if valor is None:
        return ""
    valor = str(valor)
    valor = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", valor)
    valor = valor.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    valor = re.sub(r"\s+", " ", valor)
    return valor.strip()


def extrair_peso(conteudo: str):
    texto = conteudo.upper()

    padroes = [
        r"(\d+(?:[.,]\d+)?)\s*KG",
        r"KG\s*(\d+(?:[.,]\d+)?)",
    ]

    for padrao in padroes:
        matches = re.findall(padrao, texto, flags=re.IGNORECASE)
        if matches:
            try:
                return float(matches[-1].replace(",", "."))
            except Exception:
                pass

    return None


def extrair_validade(conteudo: str) -> str:
    texto = conteudo.upper()

    padroes = [
        r"DATA\s+DE\s+VALIDADE[:\s]*([0-3]?\d/[0-1]?\d/\d{4})",
        r"VALIDADE[:\s]*([0-3]?\d/[0-1]?\d/\d{4})",
    ]

    for padrao in padroes:
        m = re.search(padrao, texto, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip()

    linhas_fd = re.findall(r"\^FD(.*?)\^FS", conteudo, flags=re.IGNORECASE | re.DOTALL)
    for linha in linhas_fd:
        linha_limpa = linha.strip()
        m = re.search(r"([0-3]?\d/[0-1]?\d/\d{4})", linha_limpa)
        if "VALIDADE" in linha_limpa.upper() and m:
            return m.group(1).strip()

    return ""


def extrair_qtd_etiquetas(conteudo: str) -> int:
    m = re.search(r"\^PQ\s*(\d+)", conteudo, flags=re.IGNORECASE)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            pass
    return 1


def extrair_descricao(conteudo: str) -> str:
    linhas_fd = re.findall(r"\^FD(.*?)\^FS", conteudo, flags=re.IGNORECASE | re.DOTALL)

    candidatos = []
    for linha in linhas_fd:
        texto = linha.strip()

        if not texto:
            continue
        if "DATA DE VALIDADE" in texto.upper():
            continue
        if "VALIDADE" in texto.upper():
            continue
        if "COD:" in texto.upper():
            continue
        if re.search(r"\b\d+(?:[.,]\d+)?\s*KG\b", texto, flags=re.IGNORECASE):
            continue
        if texto.startswith("LA,"):
            continue
        if len(texto) >= 4:
            candidatos.append(texto)

    return candidatos[0] if candidatos else ""


def copiar_e_ler_arquivo(cfg: dict, origem: Path, destino: Path):
    ultimo_erro = None

    for tentativa in range(1, MAX_TENTATIVAS + 1):
        try:
            dados = origem.read_bytes()
            destino.write_bytes(dados)
            conteudo = dados.decode("latin1", errors="ignore")
            return dados, conteudo
        except PermissionError as e:
            ultimo_erro = e
            log(cfg, f"Tentativa {tentativa}/{MAX_TENTATIVAS}: arquivo em uso -> {origem.name}")
            time.sleep(ESPERA_ENTRE_TENTATIVAS)
        except Exception as e:
            ultimo_erro = e
            log(cfg, f"Tentativa {tentativa}/{MAX_TENTATIVAS}: erro -> {e}")
            time.sleep(ESPERA_ENTRE_TENTATIVAS)

    raise ultimo_erro if ultimo_erro else RuntimeError("Falha ao acessar o arquivo.")


def montar_registro(conteudo: str, peso_extraido) -> dict:
    qtd_etiquetas = extrair_qtd_etiquetas(conteudo)
    peso_total = peso_extraido * qtd_etiquetas if peso_extraido is not None else None

    return {
        "DataHora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Descricao": limpar_texto_excel(extrair_descricao(conteudo)),
        "PesoKG": peso_extraido,
        "QtdEtiq": qtd_etiquetas,
        "PesoTotalKG": peso_total,
        "Validade": extrair_validade(conteudo),
    }


def gerar_nome_unico_fila(arquivo: Path, sessao_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    ext = arquivo.suffix.lower() if arquivo.suffix else ".txt"
    return f"{sessao_id}__{timestamp}__{uuid.uuid4().hex[:8]}{ext}"


def extrair_sessao_id_do_nome(nome_arquivo: str) -> str | None:
    partes = nome_arquivo.split("__", 2)
    if len(partes) >= 3 and partes[0]:
        return partes[0]
    return None


def mover_fila_para_descarte(cfg: dict, caminho: Path, motivo: str) -> None:
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        destino = cfg["pasta_descarte_fila"] / f"{caminho.stem}__{timestamp}{caminho.suffix}"
        shutil.move(str(caminho), str(destino))
        log(cfg, f"Arquivo movido da fila para descarte ({motivo}): {caminho.name} -> {destino.name}")
    except Exception as e:
        log(cfg, f"Falha ao mover arquivo da fila para descarte {caminho.name}: {e}")


def limpar_fila_invalida(cfg: dict) -> None:
    status = carregar_status(cfg)
    sessao_atual = status.get("sessao_id") if status.get("sessao_ativa") else None

    for caminho in cfg["pasta_fila"].iterdir():
        if not caminho.is_file():
            continue
        if caminho.suffix.lower() not in EXTENSOES_VALIDAS:
            continue

        sessao_arquivo = extrair_sessao_id_do_nome(caminho.name)

        if not sessao_atual:
            mover_fila_para_descarte(cfg, caminho, "sem sessao ativa")
            continue

        if sessao_arquivo != sessao_atual:
            mover_fila_para_descarte(cfg, caminho, "sessao diferente da atual")


def atualizar_descricao_produto_sessao(cfg: dict, descricao: str) -> None:
    if not descricao:
        return
    try:
        status = carregar_status(cfg)
        if status.get("sessao_ativa") and not status.get("descricao_produto"):
            status["descricao_produto"] = descricao
            salvar_status(cfg, status)
            log(cfg, f"Descrição do produto definida automaticamente: {descricao}")
    except Exception as e:
        log(cfg, f"Falha ao atualizar descrição do produto: {e}")


def limpar_cache_eventos_antigos(agora: float) -> None:
    limite = agora - max(JANELA_EVENTO_SEGUNDOS, LIMPEZA_CACHE_EVENTOS_SEGUNDOS)
    chaves_antigas = [
        chave for chave, ts in arquivos_capturados_recentemente.items()
        if ts < limite
    ]
    for chave in chaves_antigas:
        arquivos_capturados_recentemente.pop(chave, None)


def montar_chave_evento(cfg: dict, caminho_origem: Path):
    try:
        stat = caminho_origem.stat()
        return (
            cfg["id"],
            str(caminho_origem.resolve()),
            stat.st_size,
            int(stat.st_mtime_ns),
        )
    except Exception:
        return (
            cfg["id"],
            str(caminho_origem.resolve()),
            None,
            None,
        )


def capturar_para_fila(cfg: dict, caminho_origem: Path):
    if not caminho_origem.exists() or not caminho_origem.is_file():
        return

    status = carregar_status(cfg)
    if not status.get("sessao_ativa"):
        try:
            caminho_origem.unlink()
            log(cfg, f"Arquivo descartado porque a sessão está inativa: {caminho_origem.name}")
        except Exception as e:
            log(cfg, f"Falha ao descartar arquivo com sessão inativa {caminho_origem.name}: {e}")
        return

    sessao_id = status.get("sessao_id")
    if not sessao_id:
        try:
            caminho_origem.unlink()
            log(cfg, f"Arquivo descartado porque não havia sessao_id válido: {caminho_origem.name}")
        except Exception as e:
            log(cfg, f"Falha ao descartar arquivo sem sessao_id {caminho_origem.name}: {e}")
        return

    agora = time.time()
    chave_evento = montar_chave_evento(cfg, caminho_origem)

    with fila_lock:
        limpar_cache_eventos_antigos(agora)
        ultimo = arquivos_capturados_recentemente.get(chave_evento, 0)
        if (agora - ultimo) < JANELA_EVENTO_SEGUNDOS:
            return
        arquivos_capturados_recentemente[chave_evento] = agora

    try:
        dados = None
        for _ in range(10):
            try:
                dados = caminho_origem.read_bytes()
                if dados:
                    break
            except PermissionError:
                time.sleep(0.5)

        if not dados:
            log(cfg, f"Falha ao acessar arquivo após várias tentativas: {caminho_origem.name}")
            return

        nome_fila = gerar_nome_unico_fila(caminho_origem, sessao_id)
        destino_fila = cfg["pasta_fila"] / nome_fila
        destino_fila.write_bytes(dados)

        try:
            caminho_origem.unlink()
            log(cfg, f"Arquivo capturado para fila e removido da entrada: {caminho_origem.name} -> {destino_fila.name}")
        except Exception as e:
            log(cfg, f"Arquivo copiado para fila, mas falhou ao remover da entrada {caminho_origem.name}: {e}")

    except Exception as e:
        log(cfg, f"Erro ao capturar para fila {caminho_origem.name}: {e}")


def processar_arquivo_fila(cfg: dict, caminho_fila: Path):
    if not caminho_fila.exists() or not caminho_fila.is_file():
        return

    status = carregar_status(cfg)
    if not status.get("sessao_ativa"):
        mover_fila_para_descarte(cfg, caminho_fila, "sessao inativa ao processar")
        return

    sessao_atual = status.get("sessao_id")
    sessao_arquivo = extrair_sessao_id_do_nome(caminho_fila.name)

    if not sessao_arquivo:
        mover_fila_para_descarte(cfg, caminho_fila, "arquivo sem sessao_id")
        return

    if sessao_arquivo != sessao_atual:
        mover_fila_para_descarte(cfg, caminho_fila, "arquivo de outra sessao")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    nome_historico = f"{caminho_fila.stem}_{timestamp}{caminho_fila.suffix}"
    destino_historico = cfg["pasta_historico_etiquetas"] / nome_historico

    _, conteudo = copiar_e_ler_arquivo(cfg, caminho_fila, destino_historico)

    peso_extraido = extrair_peso(conteudo)
    if peso_extraido is None:
        log(cfg, f"Ignorado (peso não encontrado): {caminho_fila.name}")
        try:
            caminho_fila.unlink()
        except Exception as e:
            log(cfg, f"Falha ao remover arquivo ignorado da fila {caminho_fila.name}: {e}")
        return

    registro = montar_registro(conteudo, peso_extraido)

    if registro.get("Descricao"):
        atualizar_descricao_produto_sessao(cfg, registro["Descricao"])

    append_registro_excel(cfg, registro)

    try:
        caminho_fila.unlink()
        log(cfg, f"Arquivo removido da fila após processamento: {caminho_fila.name}")
    except Exception as e:
        log(cfg, f"Falha ao remover arquivo da fila {caminho_fila.name}: {e}")

    log(
        cfg,
        f"Registrado | Arquivo fila: {caminho_fila.name} | "
        f"Descricao: {registro['Descricao']} | "
        f"Qtd etiquetas: {registro['QtdEtiq']} | "
        f"Peso total: {registro['PesoTotalKG']} KG"
    )


def trabalhador_fila(cfg: dict, stop_event: threading.Event):
    while not stop_event.is_set():
        try:
            arquivos_fila = sorted(
                [
                    arq for arq in cfg["pasta_fila"].iterdir()
                    if arq.is_file() and arq.suffix.lower() in EXTENSOES_VALIDAS
                ],
                key=lambda p: p.name,
            )

            if not arquivos_fila:
                stop_event.wait(1)
                continue

            for caminho_fila in arquivos_fila:
                if stop_event.is_set():
                    break
                try:
                    processar_arquivo_fila(cfg, caminho_fila)
                except Exception as e:
                    log(cfg, f"ERRO ao processar arquivo da fila {caminho_fila.name}: {e}")

        except Exception as e:
            log(cfg, f"ERRO no trabalhador da fila: {e}")

        stop_event.wait(1)


class EtiquetaHandler(FileSystemEventHandler):
    def __init__(self, cfg: dict):
        super().__init__()
        self.cfg = cfg

    def _tratar(self, src_path: str):
        caminho = Path(src_path)
        if caminho.suffix.lower() not in EXTENSOES_VALIDAS:
            return
        capturar_para_fila(self.cfg, caminho)

    def on_created(self, event):
        if not event.is_directory:
            self._tratar(event.src_path)

    def on_modified(self, event):
        if not event.is_directory:
            self._tratar(event.src_path)


def processo_ativo(pid: int) -> bool:
    try:
        resultado = subprocess.run(
            ["tasklist", "/FI", f"PID eq {pid}"],
            capture_output=True,
            text=True
        )
        return str(pid) in resultado.stdout
    except Exception:
        return False


def salvar_pid(cfg: dict) -> None:
    dados = {
        "pid": os.getpid(),
        "impressora": cfg["id"],
        "inicio": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(cfg["arquivo_pid"], "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def remover_pid(cfg: dict) -> None:
    try:
        if cfg["arquivo_pid"].exists():
            cfg["arquivo_pid"].unlink()
    except Exception:
        pass


def status_monitor(cfg: dict) -> tuple[bool, int | None]:
    if not cfg["arquivo_pid"].exists():
        return False, None

    try:
        with open(cfg["arquivo_pid"], "r", encoding="utf-8") as f:
            dados = json.load(f)
        pid = dados.get("pid")
        if pid and processo_ativo(pid):
            return True, pid
        return False, pid
    except Exception:
        return False, None


def iniciar_monitor_processo(nome_impressora: str) -> tuple[bool, str]:
    cfg = montar_config(nome_impressora)
    garantir_pastas(cfg)
    garantir_config_limpeza(cfg)

    rodando, pid = status_monitor(cfg)
    if rodando:
        return False, f"O monitor de {cfg['nome_amigavel']} já está em execução. PID: {pid}"

    python_exe = sys.executable
    script = Path(__file__).resolve()

    processo = subprocess.Popen(
        [python_exe, str(script), nome_impressora],
        creationflags=subprocess.CREATE_NEW_CONSOLE
    )

    time.sleep(1)

    if processo_ativo(processo.pid):
        return True, f"Monitor de {cfg['nome_amigavel']} iniciado com PID {processo.pid}."
    return False, f"Falha ao iniciar o monitor de {cfg['nome_amigavel']}."


def parar_monitor_processo(nome_impressora: str) -> tuple[bool, str]:
    cfg = montar_config(nome_impressora)

    if not cfg["arquivo_pid"].exists():
        return False, f"Não há PID salvo para {cfg['nome_amigavel']}."

    try:
        with open(cfg["arquivo_pid"], "r", encoding="utf-8") as f:
            dados = json.load(f)

        pid = dados.get("pid")
        if not pid:
            return False, f"PID inválido para {cfg['nome_amigavel']}."

        if processo_ativo(pid):
            subprocess.run(
                ["taskkill", "/PID", str(pid), "/F"],
                capture_output=True,
                text=True
            )

        remover_pid(cfg)
        return True, f"Monitor de {cfg['nome_amigavel']} encerrado."
    except Exception as e:
        return False, f"Erro ao encerrar monitor de {cfg['nome_amigavel']}: {e}"


def executar_monitor(nome_impressora: str):
    cfg = montar_config(nome_impressora)
    garantir_pastas(cfg)
    garantir_arquivo_status(cfg)
    garantir_excel_atual(cfg)
    garantir_config_limpeza(cfg)
    limpar_fila_invalida(cfg)
    executar_limpeza_automatica(cfg)

    if not cfg["pasta_origem"].exists():
        raise FileNotFoundError(f"Pasta de origem não encontrada: {cfg['pasta_origem']}")

    salvar_pid(cfg)

    log(cfg, f"Iniciando monitor da entrada: {cfg['pasta_origem']}")
    log(cfg, f"Pasta de fila real: {cfg['pasta_fila']}")

    stop_event = threading.Event()
    thread_worker = threading.Thread(
        target=trabalhador_fila,
        args=(cfg, stop_event),
        daemon=True
    )
    thread_worker.start()

    event_handler = EtiquetaHandler(cfg)
    observer = Observer()
    observer.schedule(event_handler, str(cfg["pasta_origem"]), recursive=False)
    observer.start()

    ultimo_ciclo_limpeza = 0.0

    try:
        while True:
            agora = time.time()
            conf = carregar_config_limpeza(cfg)
            intervalo_limpeza = conf["intervalo_limpeza_automatica_segundos"]

            if (agora - ultimo_ciclo_limpeza) >= intervalo_limpeza:
                executar_limpeza_automatica(cfg)
                ultimo_ciclo_limpeza = agora

            time.sleep(1)
    except KeyboardInterrupt:
        log(cfg, "Encerrando monitor.")
    finally:
        stop_event.set()
        observer.stop()
        observer.join(timeout=5)
        thread_worker.join(timeout=5)
        remover_pid(cfg)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        nomes = ", ".join(CONFIG_IMPRESSORAS.keys())
        raise SystemExit(
            f"Informe a impressora. Exemplo: python monitor.py impressora_1 | Opções: {nomes}"
        )

    nome_impressora = sys.argv[1]
    executar_monitor(nome_impressora)